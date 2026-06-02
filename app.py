"""
Konsole Bulk Import File Validator — Streamlit Web App
======================================================
Validates XLS/XLSX/CSV bulk import files against Konsole's exact parsing rules.
Version: 2.1.0
"""

import streamlit as st
import re
from datetime import datetime
from io import BytesIO

# ---------------------------------------------------------------------------
# Konsole theme CSS
# ---------------------------------------------------------------------------

KONSOLE_CSS = """
<style>
/* Header bar mimicking Konsole */
.konsole-header {
    background-color: #471A91;
    padding: 12px 24px;
    border-radius: 8px;
    margin-bottom: 24px;
    display: flex;
    align-items: center;
    gap: 16px;
}
.konsole-header svg {
    flex-shrink: 0;
}
.konsole-header .header-title {
    color: #ffffff;
    font-size: 20px;
    font-weight: 600;
    letter-spacing: 0.3px;
}
.konsole-header .header-subtitle {
    color: #D5D8DA;
    font-size: 13px;
    font-weight: 400;
}

/* Status banners */
.status-pass {
    background-color: #afd427;
    color: #303C48;
    padding: 12px 20px;
    border-radius: 6px;
    font-weight: 600;
    font-size: 15px;
    margin: 12px 0;
}
.status-fail {
    background-color: #e74c3c;
    color: #ffffff;
    padding: 12px 20px;
    border-radius: 6px;
    font-weight: 600;
    font-size: 15px;
    margin: 12px 0;
}
.status-warn {
    background-color: #f8c665;
    color: #303C48;
    padding: 12px 20px;
    border-radius: 6px;
    font-weight: 600;
    font-size: 15px;
    margin: 12px 0;
}

/* Metric cards */
.metric-card {
    background-color: #F2F5F8;
    border: 1px solid #D5D8DA;
    border-radius: 6px;
    padding: 14px 18px;
    text-align: center;
}
.metric-card .metric-label {
    color: #95999c;
    font-size: 12px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 4px;
}
.metric-card .metric-value {
    color: #303C48;
    font-size: 22px;
    font-weight: 700;
}
.metric-card .metric-value.alert {
    color: #e74c3c;
}

/* Error/Warning items */
.error-item {
    background-color: #fef2f2;
    border-left: 4px solid #e74c3c;
    padding: 8px 14px;
    margin: 6px 0;
    border-radius: 0 4px 4px 0;
    font-size: 13px;
    color: #303C48;
}
.warning-item {
    background-color: #fffbeb;
    border-left: 4px solid #f8c665;
    padding: 8px 14px;
    margin: 6px 0;
    border-radius: 0 4px 4px 0;
    font-size: 13px;
    color: #303C48;
}

/* Section headers */
.section-header {
    color: #471A91;
    font-size: 16px;
    font-weight: 600;
    border-bottom: 2px solid #8545C8;
    padding-bottom: 6px;
    margin: 20px 0 12px 0;
}

/* Upload area */
[data-testid="stFileUploader"] {
    border: 2px dashed #D5D8DA;
    border-radius: 8px;
    padding: 8px;
}
[data-testid="stFileUploader"]:hover {
    border-color: #8545C8;
}

/* Purple accents on buttons */
.stButton > button {
    background-color: #8545C8;
    color: white;
    border: 1px solid #8545C8;
}
.stButton > button:hover {
    background-color: #471A91;
    border: 1px solid #471A91;
    color: white;
}

/* Expander styling */
[data-testid="stExpander"] {
    border: 1px solid #D5D8DA;
    border-radius: 6px;
}
</style>
"""

KOMGO_LOGO_SVG = """<svg width="61" height="20" viewBox="0 0 61 20" fill="none" xmlns="http://www.w3.org/2000/svg">
<path fill-rule="evenodd" clip-rule="evenodd" d="M17.3657 9.8578C17.3657 7.86444 15.9115 6.39662 14.0295 6.39662C12.1476 6.39662 10.6934 7.86444 10.6934 9.8578C10.6934 11.8512 12.1476 13.319 14.0295 13.319C15.9115 13.319 17.3657 11.8512 17.3657 9.8578M8.38281 9.85772C8.38281 6.59384 10.9063 4.09644 14.0286 4.09644C17.1723 4.09644 19.6743 6.59384 19.6743 9.85772C19.6743 13.0999 17.1723 15.619 14.0286 15.619C10.9063 15.619 8.38281 13.0999 8.38281 9.85772" fill="white"/>
<path fill-rule="evenodd" clip-rule="evenodd" d="M58.3027 9.85782C58.3027 7.86446 56.8484 6.39664 54.9665 6.39664C53.0845 6.39664 51.6303 7.86446 51.6303 9.85782C51.6303 11.8512 53.0845 13.319 54.9665 13.319C56.8484 13.319 58.3027 11.8512 58.3027 9.85782M49.3203 9.85772C49.3203 6.59384 51.8438 4.09644 54.9661 4.09644C58.1098 4.09644 60.6118 6.59384 60.6118 9.85772C60.6118 13.0999 58.1098 15.619 54.9661 15.619C51.8438 15.619 49.3203 13.0999 49.3203 9.85772" fill="white"/>
<path fill-rule="evenodd" clip-rule="evenodd" d="M2.3431 9.18346L2.30966 9.22246V0H0V15.3342H2.30966V10.274L2.3431 10.3126V9.18346Z" fill="white"/>
<path fill-rule="evenodd" clip-rule="evenodd" d="M4.57801 9.7483L9.28281 4.38135H6.45996L2.99609 8.42325V11.0647L6.69518 15.3343H9.45384L4.57801 9.7483Z" fill="white"/>
<path fill-rule="evenodd" clip-rule="evenodd" d="M32.272 4.09644C30.8392 4.09644 29.7057 4.68803 29 5.82713C28.3798 4.73171 27.3533 4.09644 25.9632 4.09644C25.0134 4.09644 24.2073 4.36778 23.5742 4.91545V7.04248C24.0068 6.54587 24.6115 6.30902 25.2789 6.30902C26.5406 6.30902 27.2892 7.16328 27.2892 8.69668V15.3342H29.5988V9.20051C29.5988 7.20715 30.5611 6.30902 31.9298 6.30902C33.1916 6.30902 34.0042 7.16328 34.0042 8.69668V15.3342H36.3139V8.54329C36.3139 5.8052 34.6886 4.09644 32.272 4.09644" fill="white"/>
<path fill-rule="evenodd" clip-rule="evenodd" d="M22.921 5.6699C22.9153 5.67871 22.9091 5.68684 22.9034 5.69574V4.38135H20.5938V15.3343H22.9034V9.2006C22.9034 9.05787 22.9104 8.92232 22.921 8.79077V5.6699Z" fill="white"/>
<path fill-rule="evenodd" clip-rule="evenodd" d="M46.1298 4.38125V6.00221C46.1223 5.99121 46.114 5.98093 46.1063 5.96996V9.33727C46.1206 9.47176 46.1298 9.60831 46.1298 9.7482C46.1298 11.7196 44.6756 13.1874 42.7295 13.1874C40.7834 13.1874 39.3292 11.7196 39.3292 9.7482C39.3292 7.77659 40.7834 6.30899 42.7295 6.30899C43.8745 6.30899 44.8463 6.81959 45.4548 7.63948V5.22238C44.6777 4.50986 43.6534 4.09644 42.4087 4.09644C39.3933 4.09644 37.0195 6.61555 37.0195 9.7482C37.0195 12.8807 39.3933 15.3998 42.4087 15.3998C44.0554 15.3998 45.3172 14.677 46.1298 13.494V14.8303C46.1298 16.7361 44.9322 17.8533 42.9434 17.8533C41.0828 17.8533 40.2915 17.1085 39.8211 16.1665L37.8322 17.3495C38.7303 19.1019 40.5481 20 42.9005 20C45.7021 20 48.3967 18.3571 48.3967 14.8303V4.38125H46.1298Z" fill="white"/>
</svg>"""

# ---------------------------------------------------------------------------
# Column definitions per message type (from Java @ExcelBindByName annotations)
# ---------------------------------------------------------------------------

GTI_COLUMNS = [
    ("(GTI) Guarantor Type", True, r"(?i)^bank$", "must be 'bank'"),
    ("(GTI) Guarantor Name and Address", False, None, None),
    ("(GTI) Borrower Full Name", False, None, None),
    ("(GTI) Corporate Ref. No.", False, None, None),
    ("(GTI) Applicant Name and Address", False, None, None),
    ("(GTI) Beneficiary Name and Address", False, None, None),
    ("(GTI) Instrument Type", True, r"(?i)^(Demand guarantee|Dependent undertaking|Standby letter of credit)$",
     "must be 'Demand guarantee', 'Dependent undertaking', or 'Standby letter of credit'"),
    ("(GTI) Local Guarantee Type", False,
     r"(?i)^(judicial|payment|advance payment|bill of lading|customs|direct pay|insurance|"
     r"lease|other type|other|performance|retention|shipping|tender or bid|warranty/maintenance)$",
     "must be one of: judicial, payment, advance payment, bill of lading, customs, direct pay, "
     "insurance, lease, other type, other, performance, retention, shipping, "
     "tender or bid, warranty/maintenance (note: 'none' is rejected by Konsole)"),
    ("(GTI) Nominal Currency", False, None, None),
    ("(GTI) Nominal Amount", False, None, None),
    ("(GTI) Issue Date", False, None, None),
    ("(GTI) Expiry Date", False, None, None),
    ("(GTI) Validity Type", False, None, None),
    ("(GTI) Local Guarantor Expiry Date", False, None, None),
    ("(GTI) Form of Guarantee", False, r"(?i)^(indirect|direct)$", "must be 'indirect' or 'direct'"),
    ("(GTI) Undertaking Number", False, None, None),
    ("(GTI) Applicable Rules", False, None, None),
    ("(GTI) Guarantee Text", False, None, None),
    ("(GTI) Confirmation Indicator", False,
     r"(?i)^(confirm|without confirmation|none|may add confirmation)$",
     "must be one of: confirm, without confirmation, none, may add confirmation"),
    ("(GTI) Delivery Courier", False, None, None),
    ("(GTI) Method of Delivery to Beneficiary", False,
     r"(?i)^(COLL|Collection|COUR|Courier|MAIL|MESS|Messenger|REGM|Registered Mail|DIGITAL|OTHR|Other)$",
     "must be one of: COLL/Collection, COUR/Courier, MAIL, MESS/Messenger, REGM/Registered Mail, DIGITAL, OTHR/Other"),
    ("(GTI) Konsole ID", False, None, None),
    ("(GTI) Konsole UUID reference", False, None, None),
    ("(GTI) Expiry Condition/Event", False, None, None),
    ("(GTI) Guarantee Type Details", False, None, None),
    ("(GTI) Local Guarantee Type Details", False, None, None),
    ("(GTI) Requested Confirmation Party", False, None, None),
    ("(GTI) Automatic Extension Period", False,
     r"(?i)^(number of calendar days after latest expiry date|same date one year later|other extension clause|none|DAYS|ONEY|OTHR)$",
     "must be one of: 'number of calendar days after latest expiry date', 'same date one year later', 'other extension clause', or 'none'"),
    ("(GTI) Automatic Extension Period Details", False, None, None),
    ("(GTI) Automatic Extension Notification Period", False, None, None),
    ("(GTI) Automatic Extension Final Expiry Date", False, None, None),
    ("(GTI) Local Guarantor Name and Address", False, None, None),
    ("(GTI) Applicable Rules Details", False, None, None),
]

LCE_COLUMNS = [
    ("(LCE) Form of Documentary Credit", True, None, None),
    ("(LCE) Documentary Credit Number", False, None, None),
    ("(LCE) Corporate Ref. No.", True, None, None),
    ("(LCE) Konsole ID", True, None, None),
    ("(LCE) Applicable Rules", False, None, None),
    ("(LCE) Applicable Rules (details if other)", False, None, None),
    ("(LCE) Place of Expiry", False, None, None),
    ("(LCE) Date of Expiry", False, None, None),
    ("(LCE) Date of Issue", False, None, None),
    ("(LCE) Applicant Name, Address", True, None, None),
    ("(LCE) Beneficiary Name, Address", True, None, None),
    ("(LCE) Advising Bank Name, Address", True, None, None),
    ("(LCE) Available By", False, None, None),
    ("(LCE) Available with Name, Address", True, None, None),
    ("(LCE) Confirmation instructions", False,
     r"(?i)^(confirm|without|may add)$",
     "must be one of: confirm, without, may add"),
    ("(LCE) Confirmation Indicator By Advising Bank", False, None, None),
    ("(LCE) Nominal Amount", True, None, None),
    ("(LCE) Nominal CCY", False, None, None),
    ("(GTI) Konsole UUID reference", False, None, None),
]

LCI_COLUMNS = [
    ("(LCI) Corporate Ref. No", False, None, None),
    ("(LCI) Issuing Bank Reference No", False, None, None),
    ("(LCI) Issuing Bank Name and Address", False, None, None),
    ("(LCI) Applicant/Issue on Behalf of Name and Address", False, None, None),
    ("(LCI) Beneficiary Name and Address", False, None, None),
    ("(LCI) Credit Form", False, None, None),
    ("(LCI) Issue Date", False, None, None),
    ("(LCI) Expiry Date", False, None, None),
    ("(LCI) Place of Expiry", False, None, None),
    ("(LCI) Actual Amount", False, None, None),
    ("(LCI) Base Currency Nominal Amount", False, None, None),
    ("(LCI) Applicable Rules", False, None, None),
    ("(LCI) Other Applicable Rules", False, None, None),
    ("(LCI) Available with", False, None, None),
    ("(LCI) Settlement by / Available by", False, None, None),
    ("(LCI) Confirmation", False, None, None),
    ("(LCI) Partial Shipment", False, None, None),
    ("(LCI) Trans-shipment", False, None, None),
    ("(LCI) Bank Charges", False, None, None),
    ("(LCI) Konsole ID", False, None, None),
    ("(LCI) Konsole UUID reference", False, None, None),
]

GTR_COLUMNS = [
    ("(GTR) Bank Reference Number", False, None, None),
    ("(GTR) Corporate Ref. No", False, None, None),
    ("(GTR) Issuing Date", False, None, None),
    ("(GTR) Instrument Type", False, None, None),
    ("(GTR) Guarantee Type", False, None, None),
    ("(GTR) Applicable Rules", False, None, None),
    ("(GTR) Applicable Rules Details", False, None, None),
    ("(GTR) Validity Type", False, None, None),
    ("(GTR) Expiry Date", False, None, None),
    ("Expected Expiry Date", False, None, None),
    ("(GTR) Issuing Bank Name and Address", False, None, None),
    ("(GTR) Applicant Name and Address", False, None, None),
    ("(GTR) Beneficiary Name and Address", False, None, None),
    ("(GTR) Nominal Amount", False, None, None),
    ("(GTR) Nominal Currency", False, None, None),
    ("(GTR) Guarantee Text", False, None, None),
    ("(GTR) Confirmation Instructions", False, None, None),
    ("(GTR) Undertaking Number", False, None, None),
    ("(GTR) Guarantor Type", False, None, None),
    ("(GTR) Konsole ID", False, None, None),
    ("(GTR) Foreign or domestic", False, None, None),
    ("(GTR) Underlying Transaction Details", False, None, None),
    ("(GTR) Delivery of Undertaking", False, None, None),
    ("(GTR) Other Delivery of Undertaking", False, None, None),
    ("(GTR) Expiry Condition/Event", False, None, None),
    ("(GTR) Konsole UUID reference", False, None, None),
]

TYPE_DEFS = {
    "GTI": (GTI_COLUMNS, "BG Issued / SB Import (GTI)"),
    "LCE": (LCE_COLUMNS, "LC Export (LCE)"),
    "LCI": (LCI_COLUMNS, "LC Import (LCI)"),
    "GTR": (GTR_COLUMNS, "BG Received / SB Export (GTR)"),
}

# Java LAST_COLUMN constant (BulkImportService.java:37)
JAVA_LAST_COLUMN = 20

# Fields with @Pattern validation — empty string "" will fail validation
# but null/missing will be skipped.
PATTERN_VALIDATED_FIELDS = {
    "(GTI) Local Guarantee Type",
    "(GTI) Form of Guarantee",
    "(GTI) Confirmation Indicator",
    "(LCE) Confirmation instructions",
}

# Konsole error messages for @Pattern validation failures (from messages.properties)
PATTERN_ERROR_MESSAGES = {
    "(GTI) Local Guarantee Type": "Local guarantee type is incorrect",
    "(GTI) Form of Guarantee": "Form of guarantee is incorrect",
    "(GTI) Confirmation Indicator": "Confirmation indicator is incorrect",
    "(LCE) Confirmation instructions": "Confirmation instructions is incorrect",
}


def is_date_column(col_name):
    return "date" in col_name.lower()


def is_amount_column(col_name):
    lower = col_name.lower()
    return "nominal amount" in lower or "actual amount" in lower


def validate_date(value):
    try:
        datetime.strptime(value, "%d/%m/%Y")
        return None
    except ValueError:
        return f"invalid date format (expected dd/MM/yyyy, got '{value}')"


def validate_amount(value):
    cleaned = value.replace(",", "")
    try:
        float(cleaned)
        return None
    except ValueError:
        return f"invalid amount (expected number, got '{value}')"


def detect_type(header_str):
    for keyword in ["LCE", "GTI", "GTR", "LCI"]:
        if keyword in header_str:
            return keyword
    return None


def parse_xls(file_bytes):
    """Parse .xls file, handling encryption."""
    import xlrd
    try:
        wb = xlrd.open_workbook(file_contents=file_bytes)
    except xlrd.biffh.XLRDError as e:
        if "encrypted" in str(e).lower():
            try:
                import msoffcrypto
                f = BytesIO(file_bytes)
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password="")
                decrypted = BytesIO()
                office_file.decrypt(decrypted)
                decrypted.seek(0)
                wb = xlrd.open_workbook(file_contents=decrypted.read())
            except ImportError:
                return None, "File is encrypted and msoffcrypto is not available."
            except Exception:
                return None, "File is password-protected and cannot be decrypted."
        else:
            return None, str(e)

    sh = wb.sheet_by_index(0)
    rows = []
    for r in range(sh.nrows):
        row = []
        for c in range(sh.ncols):
            cell = sh.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                row.append(dt.strftime("%d/%m/%Y"))
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                if cell.value == int(cell.value):
                    row.append(str(int(cell.value)))
                else:
                    row.append(str(cell.value))
            else:
                row.append(str(cell.value).strip() if cell.value else "")
        rows.append(row)
    return rows, None


def parse_xlsx(file_bytes):
    """Parse .xlsx file."""
    import openpyxl
    from datetime import datetime as dt, date
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        parsed_row = []
        for c in row:
            if c is None:
                parsed_row.append("")
            elif isinstance(c, (dt, date)):
                parsed_row.append(c.strftime("%d/%m/%Y"))
            elif isinstance(c, (int, float)):
                if c == int(c):
                    parsed_row.append(str(int(c)))
                else:
                    parsed_row.append(str(c))
            else:
                parsed_row.append(str(c).strip())
        rows.append(parsed_row)
    return rows, None


def parse_csv(file_bytes):
    """Parse CSV file."""
    import csv
    from io import StringIO
    text = file_bytes.decode("utf-8-sig")
    reader = csv.reader(StringIO(text))
    return [row for row in reader], None


def validate_rows(rows):
    """Validate parsed rows. Returns (errors, warnings, info)."""
    errors = []
    warnings = []
    info = {}

    if len(rows) < 2:
        errors.append("File has fewer than 2 rows. Need at least header row + 1 data row.")
        return errors, warnings, info

    row0_str = "|".join(rows[0])
    row1_str = "|".join(rows[1]) if len(rows) > 1 else ""

    header_row_idx = None
    if detect_type(row1_str):
        header_row_idx = 1
    elif detect_type(row0_str):
        header_row_idx = 0
        warnings.append(
            "Headers are in row 1, but Konsole reads from row 2. "
            "Move headers to row 2 (leave row 1 empty or as title)."
        )
    else:
        errors.append(
            "Cannot detect message type. Headers must contain GTI, LCE, LCI, or GTR prefix."
        )
        return errors, warnings, info

    headers = rows[header_row_idx]
    header_str = "|".join(headers)
    msg_type = detect_type(header_str)

    columns, type_name = TYPE_DEFS[msg_type]
    expected_names = {col[0] for col in columns}

    info["type"] = type_name
    info["header_row"] = header_row_idx + 1
    info["expected_cols"] = len(columns)
    info["found_cols"] = len([h for h in headers if h])

    # Column count check (Java: row.getLastCellNum() < LAST_COLUMN = 20)
    non_empty = [h for h in headers if h]
    if len(non_empty) < JAVA_LAST_COLUMN:
        errors.append(
            f"**COLUMN COUNT**: Only {len(non_empty)} non-empty header columns found. "
            f"Konsole requires at least {JAVA_LAST_COLUMN} columns "
            f"(`isRowNotWellFilled: row.getLastCellNum() >= {JAVA_LAST_COLUMN}`) "
            f"or **ALL rows will be silently skipped** with no error message."
        )

    # Header checks
    found_headers = set()
    case_mismatches = []
    unknown_columns = []

    for h in headers:
        if not h:
            continue
        if h in expected_names:
            found_headers.add(h)
        else:
            match = None
            for exp in expected_names:
                if h.lower() == exp.lower():
                    match = exp
                    break
            if match:
                case_mismatches.append((h, match))
            else:
                unknown_columns.append(h)

    for wrong, correct in case_mismatches:
        errors.append(f"**Column name case mismatch**: `{wrong}` — Konsole requires exact case: `{correct}`")

    for unk in unknown_columns:
        warnings.append(f"Unknown column `{unk}` — will be ignored by Konsole.")

    for col_name, required, _, _ in columns:
        if required and col_name not in found_headers:
            is_case_issue = any(wrong.lower() == col_name.lower() for wrong, _ in case_mismatches)
            if not is_case_issue:
                errors.append(f"**Missing required column**: `{col_name}`")

    # Data rows
    data_start = header_row_idx + 1
    if data_start >= len(rows):
        warnings.append("No data rows found after header.")
        return errors, warnings, info

    col_index_map = {}
    for idx, h in enumerate(headers):
        if h and h in expected_names:
            col_index_map[h] = idx

    case_mismatch_map = {}
    for wrong, correct in case_mismatches:
        for idx, h in enumerate(headers):
            if h == wrong:
                case_mismatch_map[correct] = idx

    data_row_count = 0
    skipped_by_column_count = 0
    rows_with_errors = set()  # Track which rows have errors
    rows_ok = set()  # Track clean rows
    for row_idx in range(data_start, len(rows)):
        row = rows[row_idx]
        if all(not cell for cell in row):
            continue

        data_row_count += 1
        excel_row = row_idx + 1

        # Simulate Java isRowNotWellFilled(row)
        last_non_empty = 0
        for ci, cell in enumerate(row):
            if cell:
                last_non_empty = ci + 1
        if last_non_empty < JAVA_LAST_COLUMN:
            if skipped_by_column_count == 0:
                errors.append(
                    f"**Row {excel_row}**: only {last_non_empty} non-empty columns "
                    f"(need >= {JAVA_LAST_COLUMN}). Konsole will **SILENTLY STOP** reading "
                    f"at this row (`isRowNotWellFilled` break). All subsequent rows are lost."
                )
            skipped_by_column_count += 1
            continue

        errors_before = len(errors)

        def get_cell(col_name):
            if col_name in col_index_map:
                idx = col_index_map[col_name]
                return row[idx] if idx < len(row) else ""
            if col_name in case_mismatch_map:
                idx = case_mismatch_map[col_name]
                return row[idx] if idx < len(row) else ""
            return None

        for col_name, required, pattern, pattern_desc in columns:
            value = get_cell(col_name)
            if value is None:
                continue
            value = value.strip() if value else ""

            if required and not value:
                errors.append(f"**Row {excel_row}**, `{col_name}`: empty — this field is required")
                continue

            # @Pattern fields: empty string "" FAILS validation in Java
            if not value and col_name in PATTERN_VALIDATED_FIELDS:
                konsole_msg = PATTERN_ERROR_MESSAGES.get(col_name, "value is incorrect")
                errors.append(
                    f"**Row {excel_row}**, `{col_name}`: {konsole_msg} — "
                    f"empty value fails @Pattern validation. "
                    f"Set to a valid value or remove the cell content entirely."
                )
                continue

            if not value:
                continue

            if pattern and not re.match(pattern, value):
                konsole_msg = PATTERN_ERROR_MESSAGES.get(col_name, "")
                prefix = f"{konsole_msg} — " if konsole_msg else ""
                errors.append(f"**Row {excel_row}**, `{col_name}`: {prefix}value `{value}` is invalid — {pattern_desc}")

            if is_date_column(col_name) and not re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", value):
                errors.append(f"**Row {excel_row}**, `{col_name}`: date must be dd/MM/yyyy, got `{value}`")
            elif is_date_column(col_name):
                err = validate_date(value)
                if err:
                    errors.append(f"**Row {excel_row}**, `{col_name}`: {err}")

            if is_amount_column(col_name):
                if "," in value:
                    warnings.append(
                        f"**Row {excel_row}**, `{col_name}`: amount `{value}` has comma separators — "
                        f"must be a plain number (e.g. 43257.83). Store as Excel numeric cell, not text."
                    )
                err = validate_amount(value)
                if err:
                    errors.append(f"**Row {excel_row}**, `{col_name}`: {err}")

            if "ref" in col_name.lower() and len(value) > 35:
                warnings.append(f"**Row {excel_row}**, `{col_name}`: {len(value)} chars — will be truncated to 35.")

        # Case-sensitive instrument type check
        _validate_instrument_type(row, col_index_map, case_mismatch_map, msg_type, excel_row, errors)

        # Business rules from Confluence migration specs
        if msg_type == "GTI":
            _validate_gti_business_rules(row, col_index_map, case_mismatch_map, excel_row, errors, warnings)
        elif msg_type == "LCE":
            _validate_lce_business_rules(row, col_index_map, case_mismatch_map, excel_row, errors, warnings)
        elif msg_type == "GTR":
            _validate_gtr_business_rules(row, col_index_map, case_mismatch_map, excel_row, errors, warnings)

        # Track row status
        if len(errors) > errors_before:
            rows_with_errors.add(excel_row)
        else:
            rows_ok.add(excel_row)

    if data_row_count == 0:
        errors.append(
            "**NO DATA ROWS** found after header. Konsole will return silently with no error message."
        )

    if skipped_by_column_count > 0:
        errors.append(
            f"**SILENT SKIP**: {skipped_by_column_count} row(s) silently skipped by Konsole "
            f"due to `isRowNotWellFilled()` (fewer than {JAVA_LAST_COLUMN} non-empty columns). "
            f"Konsole breaks on the first such row — all subsequent rows are also lost."
        )

    info["data_rows"] = data_row_count
    info["skipped_rows"] = skipped_by_column_count
    info["rows_with_errors"] = sorted(rows_with_errors)
    info["rows_ok"] = sorted(rows_ok)
    return errors, warnings, info


def _validate_instrument_type(row, col_index_map, case_mismatch_map, msg_type, excel_row, errors):
    prefix = f"({msg_type})"
    instrument_col = f"{prefix} Instrument Type"

    idx = col_index_map.get(instrument_col) or case_mismatch_map.get(instrument_col)
    if idx is None or idx >= len(row):
        return
    instrument = row[idx].strip() if row[idx] else ""
    if not instrument:
        return

    valid_forms = {"Demand guarantee", "Dependent undertaking", "Standby letter of credit"}
    if instrument not in valid_forms:
        for v in valid_forms:
            if instrument.lower() == v.lower() and instrument != v:
                errors.append(
                    f"**Row {excel_row}**, `{instrument_col}`: value `{instrument}` has wrong case. "
                    f"`mapFormOfUndertaking()` is **CASE-SENSITIVE**. Must be exactly `{v}`"
                )


def _validate_gti_business_rules(row, col_index_map, case_mismatch_map, excel_row, errors, warnings):
    """Business rules from Confluence migration spec for GTI."""

    def get(col_name):
        idx = col_index_map.get(col_name) or case_mismatch_map.get(col_name)
        if idx is not None and idx < len(row):
            return row[idx].strip() if row[idx] else ""
        return ""

    # Local Guarantee Type vs Form of Guarantee cross-validation
    local_guar_type = get("(GTI) Local Guarantee Type")
    form_of_guar = get("(GTI) Form of Guarantee")

    if local_guar_type and local_guar_type.lower() == "none":
        errors.append(
            f"**Row {excel_row}**, `(GTI) Local Guarantee Type`: Local guarantee type is incorrect — "
            f"value `NONE` is rejected by Konsole. "
            f"Leave the cell empty instead, or use a specific type (e.g. Performance, Other)."
        )
    if local_guar_type and form_of_guar and form_of_guar.lower() == "direct":
        warnings.append(
            f"**Row {excel_row}**, `(GTI) Local Guarantee Type`: value `{local_guar_type}` will be "
            f"**ignored** because Form of Guarantee is Direct (only used when Indirect)."
        )

    # Applicable Rules: must be ISPR, NONE, OTHR, UCPR, or URDG
    rules = get("(GTI) Applicable Rules")
    if rules and rules.upper() not in ("ISPR", "NONE", "OTHR", "UCPR", "URDG"):
        errors.append(
            f"**Row {excel_row}**, `(GTI) Applicable Rules`: value `{rules}` is invalid — "
            f"must be one of: ISPR, NONE, OTHR, UCPR, URDG"
        )

    # Validity Type mapping: Unlimited/Limited/Conditional
    validity = get("(GTI) Validity Type")
    if validity and validity.lower() not in ("unlimited", "limited", "conditional", "open", "fixd", "cond"):
        warnings.append(
            f"**Row {excel_row}**, `(GTI) Validity Type`: value `{validity}` may not map correctly — "
            f"expected: Unlimited, Limited, or Conditional"
        )

    # Expiry Condition/Event mandatory when Validity Type is Unlimited/Open
    if validity and validity.lower() in ("unlimited", "open"):
        expiry_cond = get("(GTI) Expiry Condition/Event")
        if not expiry_cond:
            errors.append(
                f"**Row {excel_row}**, `(GTI) Expiry Condition/Event`: empty — "
                f"**mandatory when Validity Type is Unlimited/Open**"
            )

    # Confirmation Indicator mandatory for SBLC (Standby letter of credit)
    instrument = get("(GTI) Instrument Type")
    if instrument and instrument.lower() == "standby letter of credit":
        confirm = get("(GTI) Confirmation Indicator")
        if not confirm:
            errors.append(
                f"**Row {excel_row}**, `(GTI) Confirmation Indicator`: empty — "
                f"**mandatory for Standby letter of credit (SBLC)**"
            )

    # Confirmation Indicator: "None" only valid for BGIS/DEPU, not SBLC
    confirm = get("(GTI) Confirmation Indicator")
    if confirm and confirm.lower() == "none":
        if instrument and instrument.lower() == "standby letter of credit":
            errors.append(
                f"**Row {excel_row}**, `(GTI) Confirmation Indicator`: value `None` is "
                f"**not accepted for SBLC** — only valid for BG Issued (Demand guarantee/Dependent undertaking)"
            )

    # Address validation: max 4 lines, 35 chars per line, min 2 lines
    for addr_col in ["(GTI) Applicant Name and Address", "(GTI) Beneficiary Name and Address",
                      "(GTI) Guarantor Name and Address"]:
        addr = get(addr_col)
        if addr:
            lines = addr.replace("\r\n", "\n").replace("\r", "\n").split("\n")
            non_empty_lines = [l for l in lines if l.strip()]
            if len(non_empty_lines) < 2:
                warnings.append(
                    f"**Row {excel_row}**, `{addr_col}`: only {len(non_empty_lines)} line(s) — "
                    f"minimum 2 lines required (name + address)"
                )
            if len(lines) > 4:
                errors.append(
                    f"**Row {excel_row}**, `{addr_col}`: {len(lines)} lines — maximum 4 lines allowed"
                )
            for i, line in enumerate(lines):
                if len(line) > 35:
                    warnings.append(
                        f"**Row {excel_row}**, `{addr_col}` line {i+1}: {len(line)} chars — max 35 per line"
                    )

    # Beneficiary: lines cannot end with period or special character
    bene = get("(GTI) Beneficiary Name and Address")
    if bene:
        for i, line in enumerate(bene.replace("\r\n", "\n").split("\n")):
            if line and re.search(r"[.\-!@#$%^&*()+=;:,<>?/\\|{}[\]~`]$", line.rstrip()):
                warnings.append(
                    f"**Row {excel_row}**, `(GTI) Beneficiary Name and Address` line {i+1}: "
                    f"ends with special character — Konsole may reject this"
                )

    # Currency code: 3 uppercase letters
    currency = get("(GTI) Nominal Currency")
    if currency and not re.match(r"^[A-Z]{3}$", currency):
        warnings.append(
            f"**Row {excel_row}**, `(GTI) Nominal Currency`: `{currency}` — "
            f"expected 3-letter code (e.g. EUR, USD, CHF)"
        )

    # Applicable Rules Details: mandatory when Applicable Rules is OTHR, max 35 chars
    if rules and rules.upper() == "OTHR":
        rules_details = get("(GTI) Applicable Rules Details")
        if not rules_details:
            errors.append(
                f"**Row {excel_row}**, `(GTI) Applicable Rules Details`: empty — "
                f"**mandatory when Applicable Rules is OTHR**"
            )
    rules_details = get("(GTI) Applicable Rules Details")
    if rules_details and len(rules_details) > 35:
        errors.append(
            f"**Row {excel_row}**, `(GTI) Applicable Rules Details`: {len(rules_details)} chars — maximum 35 allowed"
        )

    # Undertaking Number: max 16 chars
    undertaking = get("(GTI) Undertaking Number")
    if undertaking and len(undertaking) > 16:
        errors.append(
            f"**Row {excel_row}**, `(GTI) Undertaking Number`: {len(undertaking)} chars — maximum 16 allowed"
        )

    # Corporate Ref. No.: truncated to 35 chars
    corp_ref = get("(GTI) Corporate Ref. No.")
    if corp_ref and len(corp_ref) > 35:
        warnings.append(
            f"**Row {excel_row}**, `(GTI) Corporate Ref. No.`: {len(corp_ref)} chars — will be truncated to 35"
        )

    # Expiry Condition/Event: max 780 chars, 65 chars/line
    expiry_cond_text = get("(GTI) Expiry Condition/Event")
    if expiry_cond_text:
        if len(expiry_cond_text) > 780:
            errors.append(
                f"**Row {excel_row}**, `(GTI) Expiry Condition/Event`: {len(expiry_cond_text)} chars — maximum 780 allowed"
            )
        for i, line in enumerate(expiry_cond_text.replace("\r\n", "\n").split("\n")):
            if len(line) > 65:
                warnings.append(
                    f"**Row {excel_row}**, `(GTI) Expiry Condition/Event` line {i+1}: {len(line)} chars — max 65 per line"
                )

    # Guarantee Type Details: max 140 chars
    gt_details = get("(GTI) Guarantee Type Details")
    if gt_details and len(gt_details) > 140:
        errors.append(
            f"**Row {excel_row}**, `(GTI) Guarantee Type Details`: {len(gt_details)} chars — maximum 140 allowed"
        )

    # Local Guarantee Type Details: max 35 chars
    lgt_details = get("(GTI) Local Guarantee Type Details")
    if lgt_details and len(lgt_details) > 35:
        errors.append(
            f"**Row {excel_row}**, `(GTI) Local Guarantee Type Details`: {len(lgt_details)} chars — maximum 35 allowed"
        )

    # Automatic Extension Period Details: max 35 chars, mandatory if period = OTHR
    auto_ext_period = get("(GTI) Automatic Extension Period")
    auto_ext_details = get("(GTI) Automatic Extension Period Details")
    if auto_ext_period and auto_ext_period.lower() in ("other extension clause", "othr"):
        if not auto_ext_details:
            errors.append(
                f"**Row {excel_row}**, `(GTI) Automatic Extension Period Details`: empty — "
                f"**mandatory when Automatic Extension Period is OTHR**"
            )
    if auto_ext_details and len(auto_ext_details) > 35:
        errors.append(
            f"**Row {excel_row}**, `(GTI) Automatic Extension Period Details`: {len(auto_ext_details)} chars — maximum 35 allowed"
        )

    # Automatic Extension Notification Period: must be positive integer
    auto_ext_notif = get("(GTI) Automatic Extension Notification Period")
    if auto_ext_notif:
        if not re.match(r"^\d+$", auto_ext_notif) or int(auto_ext_notif) <= 0:
            errors.append(
                f"**Row {excel_row}**, `(GTI) Automatic Extension Notification Period`: "
                f"value `{auto_ext_notif}` is invalid — must be a positive integer (e.g. 30, 60, 90)"
            )

    # Delivery Courier: max 35 chars
    courier = get("(GTI) Delivery Courier")
    if courier and len(courier) > 35:
        errors.append(
            f"**Row {excel_row}**, `(GTI) Delivery Courier`: {len(courier)} chars — maximum 35 allowed"
        )

    # Requested Confirmation Party: address format (4 lines max, 35 chars/line)
    req_conf = get("(GTI) Requested Confirmation Party")
    if req_conf:
        lines = req_conf.replace("\r\n", "\n").replace("\r", "\n").split("\n")
        if len(lines) > 4:
            errors.append(
                f"**Row {excel_row}**, `(GTI) Requested Confirmation Party`: {len(lines)} lines — maximum 4 lines allowed"
            )
        for i, line in enumerate(lines):
            if len(line) > 35:
                warnings.append(
                    f"**Row {excel_row}**, `(GTI) Requested Confirmation Party` line {i+1}: {len(line)} chars — max 35 per line"
                )

    # Local Guarantor Name and Address: same rules as other address fields
    local_guarantor = get("(GTI) Local Guarantor Name and Address")
    if local_guarantor:
        lines = local_guarantor.replace("\r\n", "\n").replace("\r", "\n").split("\n")
        non_empty_lines = [l for l in lines if l.strip()]
        if len(non_empty_lines) < 2:
            warnings.append(
                f"**Row {excel_row}**, `(GTI) Local Guarantor Name and Address`: only {len(non_empty_lines)} line(s) — "
                f"minimum 2 lines required (name + address)"
            )
        if len(lines) > 4:
            errors.append(
                f"**Row {excel_row}**, `(GTI) Local Guarantor Name and Address`: {len(lines)} lines — maximum 4 lines allowed"
            )
        for i, line in enumerate(lines):
            if len(line) > 35:
                warnings.append(
                    f"**Row {excel_row}**, `(GTI) Local Guarantor Name and Address` line {i+1}: {len(line)} chars — max 35 per line"
                )

    # Guarantee Text: max 78,000 chars, 1,200 lines, 65 chars/line
    text = get("(GTI) Guarantee Text")
    if text:
        if len(text) > 78000:
            errors.append(
                f"**Row {excel_row}**, `(GTI) Guarantee Text`: {len(text)} chars — maximum 78,000 allowed"
            )
        text_lines = text.split("\n")
        if len(text_lines) > 1200:
            errors.append(
                f"**Row {excel_row}**, `(GTI) Guarantee Text`: {len(text_lines)} lines — maximum 1,200 allowed"
            )
        for i, line in enumerate(text_lines):
            if len(line) > 65:
                warnings.append(
                    f"**Row {excel_row}**, `(GTI) Guarantee Text` line {i+1}: {len(line)} chars — max 65 per line"
                )
                break  # Only warn once to avoid flooding


def _validate_lce_business_rules(row, col_index_map, case_mismatch_map, excel_row, errors, warnings):
    """Business rules from Confluence migration spec for LCE."""

    def get(col_name):
        idx = col_index_map.get(col_name) or case_mismatch_map.get(col_name)
        if idx is not None and idx < len(row):
            return row[idx].strip() if row[idx] else ""
        return ""

    # Form of Documentary Credit
    form = get("(LCE) Form of Documentary Credit")
    if form:
        valid = {"irrevocable", "irrevoc trans standby", "revocable", "revoc trans standby"}
        if form.lower() not in valid:
            errors.append(
                f"**Row {excel_row}**, `(LCE) Form of Documentary Credit`: value `{form}` is invalid — "
                f"must be one of: IRREVOCABLE, IRREVOC TRANS STANDBY, REVOCABLE, REVOC TRANS STANDBY"
            )

    # Documentary Credit Number: max 16 chars
    dcn = get("(LCE) Documentary Credit Number")
    if dcn and len(dcn) > 16:
        errors.append(
            f"**Row {excel_row}**, `(LCE) Documentary Credit Number`: {len(dcn)} chars — maximum 16 allowed"
        )

    # Applicable Rules
    rules = get("(LCE) Applicable Rules")
    if rules:
        valid_rules = {"ucp latest version", "othr", "ucpurr latest version",
                       "eucp latest version", "eucpurr latest version"}
        if rules.lower() not in valid_rules:
            errors.append(
                f"**Row {excel_row}**, `(LCE) Applicable Rules`: value `{rules}` is invalid — "
                f"must be one of: UCP LATEST VERSION, OTHR, UCPURR LATEST VERSION, "
                f"EUCP LATEST VERSION, EUCPURR LATEST VERSION (ISP not supported)"
            )

    # Available By
    avail = get("(LCE) Available By")
    if avail:
        valid_avail = {"by acceptance", "by negotiation", "by def payment",
                       "by payment", "by mixed payment"}
        if avail.lower() not in valid_avail:
            errors.append(
                f"**Row {excel_row}**, `(LCE) Available By`: value `{avail}` is invalid — "
                f"must be one of: BY ACCEPTANCE, BY NEGOTIATION, BY DEF PAYMENT, BY PAYMENT, BY MIXED PAYMENT"
            )

    # Confirmation Indicator By Advising Bank
    conf_ind = get("(LCE) Confirmation Indicator By Advising Bank")
    if conf_ind:
        if conf_ind.lower() not in ("without", "confirm"):
            errors.append(
                f"**Row {excel_row}**, `(LCE) Confirmation Indicator By Advising Bank`: "
                f"value `{conf_ind}` is invalid — must be WITHOUT or CONFIRM"
            )

    # Address validation for LCE
    for addr_col in ["(LCE) Applicant Name, Address", "(LCE) Beneficiary Name, Address",
                      "(LCE) Advising Bank Name, Address", "(LCE) Available with Name, Address"]:
        addr = get(addr_col)
        if addr:
            lines = addr.replace("\r\n", "\n").replace("\r", "\n").split("\n")
            non_empty_lines = [l for l in lines if l.strip()]
            if len(non_empty_lines) < 2:
                warnings.append(
                    f"**Row {excel_row}**, `{addr_col}`: only {len(non_empty_lines)} line(s) — "
                    f"minimum 2 lines required (name + address)"
                )
            if len(lines) > 4:
                errors.append(
                    f"**Row {excel_row}**, `{addr_col}`: {len(lines)} lines — maximum 4 lines allowed"
                )

    # Nominal CCY: 3-letter code
    ccy = get("(LCE) Nominal CCY")
    if ccy and not re.match(r"^[A-Z]{3}$", ccy):
        warnings.append(
            f"**Row {excel_row}**, `(LCE) Nominal CCY`: `{ccy}` — "
            f"expected 3-letter currency code (e.g. EUR, USD, CHF)"
        )

    # Corporate Ref: max 36 chars
    ref = get("(LCE) Corporate Ref. No.")
    if ref and len(ref) > 36:
        warnings.append(
            f"**Row {excel_row}**, `(LCE) Corporate Ref. No.`: {len(ref)} chars — maximum 36 allowed"
        )

    # Applicable Rules details: max 35 chars
    details = get("(LCE) Applicable Rules (details if other)")
    if details and len(details) > 35:
        errors.append(
            f"**Row {excel_row}**, `(LCE) Applicable Rules (details if other)`: "
            f"{len(details)} chars — maximum 35 allowed"
        )

    # Place of Expiry: max 29 chars
    place = get("(LCE) Place of Expiry")
    if place and len(place) > 29:
        errors.append(
            f"**Row {excel_row}**, `(LCE) Place of Expiry`: {len(place)} chars — maximum 29 allowed"
        )


def _validate_gtr_business_rules(row, col_index_map, case_mismatch_map, excel_row, errors, warnings):
    """Business rules from Confluence migration spec for GTR."""

    def get(col_name):
        idx = col_index_map.get(col_name) or case_mismatch_map.get(col_name)
        if idx is not None and idx < len(row):
            return row[idx].strip() if row[idx] else ""
        return ""

    # Bank Reference Number: max 16 chars
    bank_ref = get("(GTR) Bank Reference Number")
    if bank_ref and len(bank_ref) > 16:
        errors.append(
            f"**Row {excel_row}**, `(GTR) Bank Reference Number`: {len(bank_ref)} chars — maximum 16 allowed"
        )

    # Address validation
    for addr_col in ["(GTR) Issuing Bank Name and Address", "(GTR) Applicant Name and Address",
                      "(GTR) Beneficiary Name and Address"]:
        addr = get(addr_col)
        if addr:
            lines = addr.replace("\r\n", "\n").replace("\r", "\n").split("\n")
            non_empty_lines = [l for l in lines if l.strip()]
            if len(non_empty_lines) < 2:
                warnings.append(
                    f"**Row {excel_row}**, `{addr_col}`: only {len(non_empty_lines)} line(s) — "
                    f"minimum 2 lines required (name + address)"
                )
            if len(lines) > 4:
                errors.append(
                    f"**Row {excel_row}**, `{addr_col}`: {len(lines)} lines — maximum 4 lines allowed"
                )
            for i, line in enumerate(lines):
                if len(line) > 35:
                    warnings.append(
                        f"**Row {excel_row}**, `{addr_col}` line {i+1}: {len(line)} chars — max 35 per line"
                    )

    # Currency: 3-letter code
    currency = get("(GTR) Nominal Currency")
    if currency and not re.match(r"^[A-Z]{3}$", currency):
        warnings.append(
            f"**Row {excel_row}**, `(GTR) Nominal Currency`: `{currency}` — "
            f"expected 3-letter code (e.g. EUR, USD, CHF)"
        )


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

st.set_page_config(page_title="Konsole Bulk Import Validator", page_icon="https://www.komgo.io/favicon.ico", layout="wide")

# Inject Konsole CSS
st.markdown(KONSOLE_CSS, unsafe_allow_html=True)

# Header bar with komgo logo
st.markdown(f"""
<div class="konsole-header">
    {KOMGO_LOGO_SVG}
    <div>
        <div class="header-title">Bulk Import Validator</div>
        <div class="header-subtitle">Validate files against Konsole's exact parsing rules before importing</div>
    </div>
</div>
""", unsafe_allow_html=True)

uploaded = st.file_uploader("Upload bulk import file", type=["xls", "xlsx", "csv"], label_visibility="collapsed")

if uploaded:
    file_bytes = uploaded.read()
    filename = uploaded.name.lower()

    with st.spinner("Validating..."):
        rows = None
        parse_error = None

        if filename.endswith(".xls"):
            rows, parse_error = parse_xls(file_bytes)
        elif filename.endswith(".xlsx"):
            rows, parse_error = parse_xlsx(file_bytes)
        elif filename.endswith(".csv"):
            rows, parse_error = parse_csv(file_bytes)
        else:
            parse_error = "Unsupported file format."

        if parse_error:
            st.error(f"Could not read file: {parse_error}")
        elif rows:
            errors, warnings, info = validate_rows(rows)

            # Metrics row
            if info:
                skipped = info.get("skipped_rows", 0)
                if skipped > 0:
                    c1, c2, c3, c4, c5 = st.columns(5)
                else:
                    c1, c2, c3, c4 = st.columns(4)

                with c1:
                    st.markdown(f"""<div class="metric-card">
                        <div class="metric-label">Message Type</div>
                        <div class="metric-value">{info.get("type", "Unknown").split("(")[1].rstrip(")") if "(" in info.get("type", "") else info.get("type", "?")}</div>
                    </div>""", unsafe_allow_html=True)
                with c2:
                    st.markdown(f"""<div class="metric-card">
                        <div class="metric-label">Columns Found</div>
                        <div class="metric-value">{info.get("found_cols", 0)} / {info.get("expected_cols", "?")}</div>
                    </div>""", unsafe_allow_html=True)
                with c3:
                    st.markdown(f"""<div class="metric-card">
                        <div class="metric-label">Data Rows</div>
                        <div class="metric-value">{info.get("data_rows", 0)}</div>
                    </div>""", unsafe_allow_html=True)
                with c4:
                    err_class = ' alert' if errors else ''
                    st.markdown(f"""<div class="metric-card">
                        <div class="metric-label">Issues</div>
                        <div class="metric-value{err_class}">{len(errors)} E / {len(warnings)} W</div>
                    </div>""", unsafe_allow_html=True)
                if skipped > 0:
                    with c5:
                        st.markdown(f"""<div class="metric-card">
                            <div class="metric-label">Silently Skipped</div>
                            <div class="metric-value alert">{skipped}</div>
                        </div>""", unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # Result banner
            if not errors and not warnings:
                st.markdown('<div class="status-pass">PASS — File looks valid for Konsole bulk import</div>', unsafe_allow_html=True)
            elif errors:
                st.markdown(f'<div class="status-fail">FAIL — {len(errors)} error(s), {len(warnings)} warning(s)</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="status-warn">WARNING — {len(warnings)} warning(s), review recommended</div>', unsafe_allow_html=True)

            # Row summary — which rows will be processed vs rejected
            rows_err = info.get("rows_with_errors", [])
            rows_pass = info.get("rows_ok", [])
            if rows_err or rows_pass:
                st.markdown('<div class="section-header">Row Summary</div>', unsafe_allow_html=True)

                def _format_ranges(row_list):
                    """Compress [3,4,5,8,9,12] into '3-5, 8-9, 12'."""
                    if not row_list:
                        return "—"
                    ranges = []
                    start = row_list[0]
                    end = start
                    for r in row_list[1:]:
                        if r == end + 1:
                            end = r
                        else:
                            ranges.append(f"{start}-{end}" if start != end else str(start))
                            start = end = r
                    ranges.append(f"{start}-{end}" if start != end else str(start))
                    return ", ".join(ranges)

                col_a, col_b = st.columns(2)
                with col_a:
                    if rows_pass:
                        st.markdown(
                            f'<div style="background:#f0fdf4;border-left:4px solid #afd427;padding:10px 14px;border-radius:0 4px 4px 0;margin:4px 0;">'
                            f'<strong style="color:#166534;">Will be imported ({len(rows_pass)} rows)</strong><br>'
                            f'<span style="color:#303C48;font-size:13px;">Rows: {_format_ranges(rows_pass)}</span></div>',
                            unsafe_allow_html=True)
                    else:
                        st.markdown(
                            '<div style="background:#fef2f2;padding:10px 14px;border-radius:4px;color:#991b1b;">'
                            '<strong>No rows will be imported</strong></div>',
                            unsafe_allow_html=True)
                with col_b:
                    if rows_err:
                        st.markdown(
                            f'<div style="background:#fef2f2;border-left:4px solid #e74c3c;padding:10px 14px;border-radius:0 4px 4px 0;margin:4px 0;">'
                            f'<strong style="color:#991b1b;">Will be rejected ({len(rows_err)} rows)</strong><br>'
                            f'<span style="color:#303C48;font-size:13px;">Rows: {_format_ranges(rows_err)}</span></div>',
                            unsafe_allow_html=True)

            # Errors grouped by row
            if errors:
                st.markdown(f'<div class="section-header">Errors ({len(errors)})</div>', unsafe_allow_html=True)

                # Group errors by row number
                import re as _re
                grouped = {}
                general_errors = []
                for e in errors:
                    m = _re.match(r"\*\*Row (\d+)\*\*", e)
                    if m:
                        row_num = int(m.group(1))
                        grouped.setdefault(row_num, []).append(e)
                    else:
                        general_errors.append(e)

                # Show general errors first
                for e in general_errors:
                    st.markdown(f'<div class="error-item">{e}</div>', unsafe_allow_html=True)

                # Show per-row errors in expandable sections
                for row_num in sorted(grouped.keys()):
                    row_errors = grouped[row_num]
                    with st.expander(f"Row {row_num} — {len(row_errors)} error(s)", expanded=False):
                        for e in row_errors:
                            # Remove the "**Row N**, " prefix for cleaner display inside the group
                            clean = _re.sub(r"^\*\*Row \d+\*\*, ", "", e)
                            st.markdown(f'<div class="error-item">{clean}</div>', unsafe_allow_html=True)

            # Warnings grouped by row
            if warnings:
                st.markdown(f'<div class="section-header">Warnings ({len(warnings)})</div>', unsafe_allow_html=True)

                import re as _re
                grouped_w = {}
                general_warnings = []
                for w in warnings:
                    m = _re.match(r"\*\*Row (\d+)\*\*", w)
                    if m:
                        row_num = int(m.group(1))
                        grouped_w.setdefault(row_num, []).append(w)
                    else:
                        general_warnings.append(w)

                for w in general_warnings:
                    st.markdown(f'<div class="warning-item">{w}</div>', unsafe_allow_html=True)

                if grouped_w:
                    with st.expander(f"Row-level warnings ({sum(len(v) for v in grouped_w.values())})", expanded=False):
                        for row_num in sorted(grouped_w.keys()):
                            for w in grouped_w[row_num]:
                                clean = _re.sub(r"^\*\*Row \d+\*\*, ", "", w)
                                st.markdown(f'<div class="warning-item">**Row {row_num}** — {clean}</div>', unsafe_allow_html=True)

            # Reference section
            with st.expander("Accepted column names reference"):
                for type_key, (cols_def, type_name) in TYPE_DEFS.items():
                    st.markdown(f"**{type_name}**")
                    for col_name, required, pattern, desc in cols_def:
                        req_tag = " *(required)*" if required else ""
                        constraint = f" — {desc}" if desc else ""
                        st.markdown(f"- `{col_name}`{req_tag}{constraint}")
                    st.markdown("")

            with st.expander("Instrument Type — case-sensitive values"):
                st.markdown(
                    "The `mapFormOfUndertaking()` function uses a **case-sensitive** switch. "
                    "Only these exact values are accepted:\n\n"
                    "| Value (exact case) | Maps to |\n"
                    "|---|---|\n"
                    "| `Demand guarantee` | DGAR |\n"
                    "| `Dependent undertaking` | DEPU |\n"
                    "| `Standby letter of credit` | STBY |\n\n"
                    "Any other casing (e.g. `Standby Letter of Credit`, `DEMAND GUARANTEE`) will fail silently."
                )
else:
    # Empty state
    st.markdown("""
    <div style="text-align: center; padding: 60px 20px; color: #95999c;">
        <div style="font-size: 48px; margin-bottom: 16px;">&#128196;</div>
        <div style="font-size: 16px; font-weight: 500; margin-bottom: 8px;">Drop your bulk import file here</div>
        <div style="font-size: 13px;">Supports .xls, .xlsx, and .csv files</div>
    </div>
    """, unsafe_allow_html=True)
